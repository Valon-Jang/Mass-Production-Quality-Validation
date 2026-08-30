"""Read-only OOXML workbook scanner backed by openpyxl."""

from __future__ import annotations

import hashlib
import re
from dataclasses import replace
from pathlib import Path
from typing import Any, BinaryIO, Never
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile, ZipInfo

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.chartsheet import Chartsheet  # type: ignore[import-untyped]
from openpyxl.utils import (  # type: ignore[import-untyped]
    column_index_from_string,
    get_column_letter,
)
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore[import-untyped]
from openpyxl.utils.units import EMU_to_pixels  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from app.domain.workbook_scan import (
    CellEvidence,
    DisplayValueStatus,
    ImageMetadata,
    IndexRange,
    IssueSeverity,
    MacroHandling,
    RowCandidate,
    RowCandidateKind,
    ScanIssue,
    ScanPolicy,
    SheetKind,
    SheetProtectionMetadata,
    SheetScan,
    SourceLocation,
    WorkbookScan,
    WorkbookScanFailure,
    WorkbookScanFailureStatus,
    WorkbookScanState,
)

_ALLOWED_SUFFIXES = frozenset({".xlsx", ".xlsm"})
_OLE_COMPOUND_HEADER = bytes.fromhex("D0CF11E0A1B11AE1")
_EXTERNAL_FORMULA_PATTERN = re.compile(r"\[[^\]]+\][^!]*!", re.IGNORECASE)
_CONTENT_TYPES_PART = "[Content_Types].xml"
_WORKBOOK_PART = "xl/workbook.xml"
_XML_PART_SUFFIXES = (".xml", ".rels")
_WORKBOOK_MAIN_CONTENT_TYPES = {
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
    ".xlsm": "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
}
_PROTECTION_ACTIONS = (
    "autoFilter",
    "deleteColumns",
    "deleteRows",
    "formatCells",
    "formatColumns",
    "formatRows",
    "insertColumns",
    "insertHyperlinks",
    "insertRows",
    "objects",
    "pivotTables",
    "scenarios",
    "selectLockedCells",
    "selectUnlockedCells",
    "sort",
)


class OpenpyxlWorkbookScanner:
    """Scan OOXML evidence without saving, editing, unlocking, or running VBA."""

    def scan(self, source: Path, policy: ScanPolicy | None = None) -> WorkbookScan:
        active_policy = policy or ScanPolicy()
        source = Path(source)

        if not source.is_file():
            raise self._failure(
                WorkbookScanFailureStatus.FILE_NOT_FOUND,
                "WORKBOOK_FILE_NOT_FOUND",
                f"Workbook file does not exist: {source.name}",
                SourceLocation.workbook(),
            )

        before_hash = self._sha256(source)
        try:
            with source.open("rb") as source_stream:
                partial = self.scan_stream(
                    source_stream,
                    source_name=source.name,
                    policy=active_policy,
                )
        except WorkbookScanFailure as exc:
            self._raise_with_integrity_evidence(source, before_hash, exc)
        except OSError:
            failure = self._failure(
                WorkbookScanFailureStatus.WORKBOOK_UNREADABLE,
                "WORKBOOK_READ_FAILED",
                "Workbook could not be opened from the source path.",
                SourceLocation.workbook(),
            )
            self._raise_with_integrity_evidence(source, before_hash, failure)

        after_hash = self._sha256_or_none(source)
        if after_hash != before_hash:
            raise self._source_mutated_failure(before_hash, after_hash)

        return replace(
            partial,
            source_sha256_before=before_hash,
            source_sha256_after=after_hash,
        )

    def scan_stream(
        self,
        source: BinaryIO,
        *,
        source_name: str,
        policy: ScanPolicy | None = None,
    ) -> WorkbookScan:
        """Scan a File Store stream without requiring or exposing its absolute path."""

        active_policy = policy or ScanPolicy()
        logical_name = Path(source_name).name
        original_position = self._validated_stream_position(source)
        try:
            return self._scan_stream_with_integrity(
                source,
                source_name=logical_name,
                policy=active_policy,
            )
        finally:
            try:
                source.seek(original_position)
            except (OSError, ValueError) as exc:
                raise self._failure(
                    WorkbookScanFailureStatus.WORKBOOK_UNREADABLE,
                    "WORKBOOK_STREAM_POSITION_RESTORE_FAILED",
                    "Seekable source stream could not be restored to its original position.",
                    SourceLocation.workbook(),
                ) from exc

    def _scan_stream_with_integrity(
        self,
        source: BinaryIO,
        *,
        source_name: str,
        policy: ScanPolicy,
    ) -> WorkbookScan:
        before_hash, source_size = self._sha256_stream(source)
        try:
            suffix = Path(source_name).suffix.lower()
            if suffix not in _ALLOWED_SUFFIXES:
                raise self._failure(
                    WorkbookScanFailureStatus.UNSUPPORTED_FORMAT,
                    "UNSUPPORTED_WORKBOOK_FORMAT",
                    "Only .xlsx and .xlsm OOXML workbooks are supported.",
                    SourceLocation.workbook(),
                )
            self._preflight_ooxml(
                source,
                source_name=source_name,
                policy=policy,
            )
            partial = self._scan_ooxml(
                source,
                source_name=source_name,
                source_size=source_size,
                policy=policy,
            )
        except WorkbookScanFailure as exc:
            self._raise_with_stream_integrity_evidence(source, before_hash, exc)

        after_hash, _ = self._sha256_stream(source)
        if after_hash != before_hash:
            raise self._source_mutated_failure(before_hash, after_hash)
        return replace(
            partial,
            source_sha256_before=before_hash,
            source_sha256_after=after_hash,
        )

    def _preflight_ooxml(
        self,
        source: BinaryIO,
        *,
        source_name: str,
        policy: ScanPolicy,
    ) -> None:
        original_position = self._stream_position(source)
        try:
            source.seek(0)
            header = source.read(len(_OLE_COMPOUND_HEADER))
        except (OSError, ValueError) as exc:
            raise self._failure(
                WorkbookScanFailureStatus.WORKBOOK_UNREADABLE,
                "WORKBOOK_READ_FAILED",
                "Workbook could not be read from the File Store stream.",
                SourceLocation.workbook(),
            ) from exc

        if header == _OLE_COMPOUND_HEADER:
            raise self._failure(
                WorkbookScanFailureStatus.ENCRYPTED_OR_BINARY_WORKBOOK,
                "ENCRYPTED_OR_BINARY_WORKBOOK_UNSUPPORTED",
                (
                    "The file is an encrypted Office container or legacy binary workbook; "
                    "no password or protection bypass was attempted."
                ),
                SourceLocation.workbook(),
            )

        try:
            source.seek(0)
            with ZipFile(source) as archive:
                parts = tuple(info for info in archive.infolist() if not info.is_dir())
                self._reject_duplicate_package_parts(parts)
                self._enforce_package_preflight_limits(parts, policy)
                parts_by_name = {info.filename: info for info in parts}
                for required_part in (_CONTENT_TYPES_PART, _WORKBOOK_PART):
                    if required_part not in parts_by_name:
                        raise self._failure(
                            WorkbookScanFailureStatus.CORRUPT_OOXML,
                            "REQUIRED_OOXML_PART_MISSING",
                            f"Required OOXML package part is missing: {required_part}",
                            SourceLocation.part(required_part),
                        )

                for info in parts:
                    if info.flag_bits & 0x1:
                        raise self._failure(
                            WorkbookScanFailureStatus.ENCRYPTED_OR_BINARY_WORKBOOK,
                            "ENCRYPTED_OOXML_PART_UNSUPPORTED",
                            (
                                "An encrypted OOXML package part cannot be scanned without "
                                "bypassing protection."
                            ),
                            SourceLocation.part(info.filename),
                        )

                content_types_info = parts_by_name[_CONTENT_TYPES_PART]
                content_types = self._read_bounded_package_part(
                    archive,
                    content_types_info,
                    policy,
                )
                content_types_root = self._parse_xml_part(
                    content_types,
                    content_types_info.filename,
                )
                self._validate_workbook_main_content_type(
                    content_types_root,
                    suffix=Path(source_name).suffix.lower(),
                )

                for info in parts:
                    if not info.filename.lower().endswith(_XML_PART_SUFFIXES):
                        continue
                    if info.filename == _CONTENT_TYPES_PART:
                        continue
                    content = self._read_bounded_package_part(archive, info, policy)
                    self._parse_xml_part(content, info.filename)
        except WorkbookScanFailure:
            raise
        except (BadZipFile, EOFError) as exc:
            raise self._failure(
                WorkbookScanFailureStatus.CORRUPT_OOXML,
                "CORRUPT_OOXML_PACKAGE",
                "The source is not a readable OOXML ZIP package.",
                SourceLocation.workbook(),
            ) from exc
        except OSError as exc:
            raise self._failure(
                WorkbookScanFailureStatus.WORKBOOK_UNREADABLE,
                "WORKBOOK_READ_FAILED",
                "Workbook package could not be read.",
                SourceLocation.workbook(),
            ) from exc
        finally:
            try:
                source.seek(original_position)
            except (OSError, ValueError) as exc:
                raise self._failure(
                    WorkbookScanFailureStatus.WORKBOOK_UNREADABLE,
                    "WORKBOOK_STREAM_POSITION_RESTORE_FAILED",
                    "Source stream position could not be restored after OOXML preflight.",
                    SourceLocation.workbook(),
                ) from exc

    def _reject_duplicate_package_parts(self, parts: tuple[ZipInfo, ...]) -> None:
        seen: set[str] = set()
        duplicate_names: set[str] = set()
        for info in parts:
            if info.filename in seen:
                duplicate_names.add(info.filename)
            else:
                seen.add(info.filename)
        if duplicate_names:
            duplicate_name = sorted(duplicate_names)[0]
            raise self._failure(
                WorkbookScanFailureStatus.CORRUPT_OOXML,
                "DUPLICATE_OOXML_PACKAGE_PART",
                (
                    "OOXML package declares the same non-directory part more than once: "
                    f"{duplicate_name}"
                ),
                SourceLocation.part(duplicate_name),
            )

    def _enforce_package_preflight_limits(
        self,
        parts: tuple[ZipInfo, ...],
        policy: ScanPolicy,
    ) -> None:
        if len(parts) > policy.max_package_parts:
            raise self._failure(
                WorkbookScanFailureStatus.PACKAGE_PART_COUNT_LIMIT_EXCEEDED,
                "PACKAGE_PART_COUNT_LIMIT_EXCEEDED",
                (
                    f"OOXML package has {len(parts)} parts, above the provisional "
                    f"max_package_parts={policy.max_package_parts}."
                ),
                SourceLocation.workbook(),
            )

        total_uncompressed = 0
        total_xml_uncompressed = 0
        for info in parts:
            if info.file_size > policy.max_part_uncompressed_bytes:
                raise self._failure(
                    WorkbookScanFailureStatus.PACKAGE_PART_SIZE_LIMIT_EXCEEDED,
                    "PACKAGE_PART_SIZE_LIMIT_EXCEEDED",
                    (
                        f"OOXML part expands to {info.file_size} bytes, above the provisional "
                        f"max_part_uncompressed_bytes={policy.max_part_uncompressed_bytes}."
                    ),
                    SourceLocation.part(info.filename),
                )

            total_uncompressed += info.file_size
            if total_uncompressed > policy.max_total_uncompressed_bytes:
                raise self._failure(
                    WorkbookScanFailureStatus.PACKAGE_TOTAL_SIZE_LIMIT_EXCEEDED,
                    "PACKAGE_TOTAL_SIZE_LIMIT_EXCEEDED",
                    (
                        f"OOXML package expands beyond the provisional "
                        f"max_total_uncompressed_bytes={policy.max_total_uncompressed_bytes}."
                    ),
                    SourceLocation.workbook(),
                )

            if info.filename.lower().endswith(_XML_PART_SUFFIXES):
                total_xml_uncompressed += info.file_size
                if total_xml_uncompressed > policy.max_total_xml_uncompressed_bytes:
                    raise self._failure(
                        WorkbookScanFailureStatus.PACKAGE_XML_SIZE_LIMIT_EXCEEDED,
                        "PACKAGE_XML_SIZE_LIMIT_EXCEEDED",
                        (
                            "OOXML XML content expands beyond the provisional "
                            "max_total_xml_uncompressed_bytes="
                            f"{policy.max_total_xml_uncompressed_bytes}."
                        ),
                        SourceLocation.part(info.filename),
                    )

    def _read_bounded_package_part(
        self,
        archive: ZipFile,
        info: ZipInfo,
        policy: ScanPolicy,
    ) -> bytes:
        try:
            with archive.open(info) as part_stream:
                content = part_stream.read(policy.max_part_uncompressed_bytes + 1)
        except (BadZipFile, EOFError, NotImplementedError, OSError, RuntimeError) as exc:
            raise self._failure(
                WorkbookScanFailureStatus.CORRUPT_OOXML,
                "CORRUPT_OOXML_PART",
                f"OOXML package part cannot be read: {info.filename}",
                SourceLocation.part(info.filename),
            ) from exc
        if len(content) > policy.max_part_uncompressed_bytes:
            raise self._failure(
                WorkbookScanFailureStatus.PACKAGE_PART_SIZE_LIMIT_EXCEEDED,
                "PACKAGE_PART_SIZE_LIMIT_EXCEEDED",
                (
                    "OOXML part exceeded max_part_uncompressed_bytes while being read: "
                    f"{info.filename}"
                ),
                SourceLocation.part(info.filename),
            )
        return content

    def _parse_xml_part(self, content: bytes, package_part: str) -> ElementTree.Element:
        try:
            return ElementTree.fromstring(content)
        except ElementTree.ParseError as exc:
            raise self._failure(
                WorkbookScanFailureStatus.CORRUPT_OOXML,
                "CORRUPT_OOXML_PART",
                f"OOXML package part cannot be parsed: {package_part}",
                SourceLocation.part(package_part),
            ) from exc

    def _validate_workbook_main_content_type(
        self,
        content_types_root: ElementTree.Element,
        *,
        suffix: str,
    ) -> None:
        defaults: dict[str, str] = {}
        overrides: dict[str, str] = {}
        for element in content_types_root:
            declaration_kind = element.tag.rsplit("}", maxsplit=1)[-1]
            if declaration_kind == "Default":
                extension = element.attrib.get("Extension", "").lstrip(".").casefold()
                content_type = element.attrib.get("ContentType", "").strip()
                self._validate_content_type_declaration(extension, content_type, "Default")
                if extension in defaults:
                    self._raise_duplicate_content_type_declaration(
                        "Default",
                        extension,
                        defaults[extension],
                        content_type,
                    )
                defaults[extension] = content_type
            elif declaration_kind == "Override":
                part_name = element.attrib.get("PartName", "").lstrip("/").casefold()
                content_type = element.attrib.get("ContentType", "").strip()
                self._validate_content_type_declaration(part_name, content_type, "Override")
                if part_name in overrides:
                    self._raise_duplicate_content_type_declaration(
                        "Override",
                        part_name,
                        overrides[part_name],
                        content_type,
                    )
                overrides[part_name] = content_type

        workbook_part = _WORKBOOK_PART.casefold()
        declared_type = overrides.get(workbook_part)
        if declared_type is None:
            workbook_extension = Path(_WORKBOOK_PART).suffix.lstrip(".").casefold()
            declared_type = defaults.get(workbook_extension)
        if declared_type is None:
            raise self._failure(
                WorkbookScanFailureStatus.CORRUPT_OOXML,
                "WORKBOOK_MAIN_CONTENT_TYPE_MISSING",
                (
                    "No exact Override or part-extension Default resolves the main "
                    "ContentType for /xl/workbook.xml."
                ),
                SourceLocation.part(_CONTENT_TYPES_PART),
            )

        expected = _WORKBOOK_MAIN_CONTENT_TYPES[suffix]
        if declared_type.casefold() != expected.casefold():
            raise self._failure(
                WorkbookScanFailureStatus.CORRUPT_OOXML,
                "WORKBOOK_MAIN_CONTENT_TYPE_MISMATCH",
                (
                    f"Logical {suffix} extension does not match workbook main ContentType "
                    f"{declared_type!r}."
                ),
                SourceLocation.part(_CONTENT_TYPES_PART),
            )

    def _validate_content_type_declaration(
        self,
        key: str,
        content_type: str,
        declaration_kind: str,
    ) -> None:
        if key and content_type:
            return
        raise self._failure(
            WorkbookScanFailureStatus.CORRUPT_OOXML,
            "INVALID_CONTENT_TYPE_DECLARATION",
            f"{declaration_kind} content-type declaration has a blank key or ContentType.",
            SourceLocation.part(_CONTENT_TYPES_PART),
        )

    def _raise_duplicate_content_type_declaration(
        self,
        declaration_kind: str,
        key: str,
        first_content_type: str,
        second_content_type: str,
    ) -> Never:
        qualifier = (
            "conflicting"
            if first_content_type.casefold() != second_content_type.casefold()
            else "duplicate"
        )
        raise self._failure(
            WorkbookScanFailureStatus.CORRUPT_OOXML,
            f"{declaration_kind.upper()}_CONTENT_TYPE_DECLARATION_DUPLICATE",
            (
                f"[Content_Types].xml contains a {qualifier} {declaration_kind} "
                f"declaration for {key!r}."
            ),
            SourceLocation.part(_CONTENT_TYPES_PART),
        )

    def _scan_ooxml(
        self,
        source: BinaryIO,
        *,
        source_name: str,
        source_size: int,
        policy: ScanPolicy,
    ) -> WorkbookScan:
        formula_workbook = self._open_workbook(source, data_only=False)
        cached_workbook: Any | None = None
        try:
            estimated_cells = self._enforce_max_cells(formula_workbook, policy)
            cached_workbook = self._open_workbook(source, data_only=True)
            if formula_workbook.sheetnames != cached_workbook.sheetnames:
                raise self._failure(
                    WorkbookScanFailureStatus.CORRUPT_OOXML,
                    "WORKBOOK_VIEW_SHEET_MISMATCH",
                    "Formula and cached-value views expose different sheet lists.",
                    SourceLocation.workbook(),
                )

            global_issues = [
                ScanIssue(
                    code="DISPLAY_VALUE_NOT_RENDERED",
                    severity=IssueSeverity.INFO,
                    message=(
                        "Stored value, formula cache, and number format are preserved separately; "
                        "this scanner is not an Excel display-value renderer."
                    ),
                    location=SourceLocation.workbook(),
                )
            ]
            external_links = tuple(getattr(formula_workbook, "_external_links", ()))
            if external_links:
                global_issues.append(
                    ScanIssue(
                        code="EXTERNAL_LINKS_PRESENT",
                        severity=IssueSeverity.WARNING,
                        message="Workbook package contains external-link relationships.",
                        location=SourceLocation.workbook(),
                    )
                )

            macro_handling = MacroHandling.NOT_APPLICABLE
            if Path(source_name).suffix.lower() == ".xlsm":
                macro_handling = MacroHandling.NOT_LOADED_OR_EXECUTED
                global_issues.extend(
                    (
                        ScanIssue(
                            code="VBA_NOT_LOADED_OR_EXECUTED",
                            severity=IssueSeverity.WARNING,
                            message=(
                                "Macro-capable workbook was opened with keep_vba=False; "
                                "VBA was not loaded or executed."
                            ),
                            location=SourceLocation.workbook(),
                        ),
                        ScanIssue(
                            code="CALCULATION_REFRESH_REQUIRED",
                            severity=IssueSeverity.WARNING,
                            message=(
                                "Macro-dependent or stale cached results require an external "
                                "calculation refresh before official use."
                            ),
                            location=SourceLocation.workbook(),
                        ),
                    )
                )

            security = getattr(formula_workbook, "security", None)
            if security is not None and bool(getattr(security, "lockStructure", False)):
                global_issues.append(
                    ScanIssue(
                        code="WORKBOOK_STRUCTURE_PROTECTED",
                        severity=IssueSeverity.INFO,
                        message="Workbook structure protection was observed and was not bypassed.",
                        location=SourceLocation.workbook(),
                    )
                )

            sheets: list[SheetScan] = []
            for position, sheet_name in enumerate(formula_workbook.sheetnames):
                formula_sheet = formula_workbook[sheet_name]
                cached_sheet = cached_workbook[sheet_name]
                if isinstance(formula_sheet, Worksheet) and isinstance(cached_sheet, Worksheet):
                    sheets.append(
                        self._scan_worksheet(
                            formula_sheet,
                            cached_sheet,
                            position=position,
                        )
                    )
                elif isinstance(formula_sheet, Chartsheet) and isinstance(cached_sheet, Chartsheet):
                    sheets.append(self._scan_chartsheet(formula_sheet, position=position))
                else:
                    raise self._failure(
                        WorkbookScanFailureStatus.CORRUPT_OOXML,
                        "SHEET_TYPE_MISMATCH",
                        f"Formula and cached-value views disagree on sheet type: {sheet_name}",
                        SourceLocation.sheet(sheet_name),
                    )

            all_issues = tuple(global_issues) + tuple(
                issue for sheet in sheets for issue in sheet.issues
            )
            has_warnings = any(issue.severity == IssueSeverity.WARNING for issue in all_issues)
            return WorkbookScan(
                state=(
                    WorkbookScanState.SCANNED_WITH_WARNINGS
                    if has_warnings
                    else WorkbookScanState.SCANNED
                ),
                source_name=source_name,
                source_size_bytes=source_size,
                source_sha256_before="",
                source_sha256_after="",
                sheets=tuple(sheets),
                issues=all_issues,
                estimated_cells=estimated_cells,
                external_link_count=len(external_links),
                macro_handling=macro_handling,
                display_value_contract=DisplayValueStatus.NOT_RENDERED,
                is_golden_workbook_evidence=False,
            )
        finally:
            formula_workbook.close()
            if cached_workbook is not None:
                cached_workbook.close()

    def _open_workbook(self, source: BinaryIO, *, data_only: bool) -> Any:
        try:
            source.seek(0)
            return load_workbook(
                filename=source,
                data_only=data_only,
                keep_links=True,
                keep_vba=False,
                read_only=False,
            )
        except InvalidFileException as exc:
            raise self._failure(
                WorkbookScanFailureStatus.CORRUPT_OOXML,
                "OPENPYXL_REJECTED_WORKBOOK",
                "openpyxl rejected the OOXML workbook.",
                SourceLocation.workbook(),
            ) from exc
        except (BadZipFile, KeyError, ElementTree.ParseError, ValueError) as exc:
            raise self._failure(
                WorkbookScanFailureStatus.CORRUPT_OOXML,
                "WORKBOOK_PARSE_FAILED",
                "Workbook OOXML could not be parsed.",
                SourceLocation.workbook(),
            ) from exc
        except OSError as exc:
            raise self._failure(
                WorkbookScanFailureStatus.WORKBOOK_UNREADABLE,
                "WORKBOOK_READ_FAILED",
                "Workbook could not be read.",
                SourceLocation.workbook(),
            ) from exc

    def _enforce_max_cells(self, workbook: Any, policy: ScanPolicy) -> int:
        total = 0
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if not isinstance(sheet, Worksheet):
                continue
            sheet_cells = int(sheet.max_row) * int(sheet.max_column)
            total += sheet_cells
            if total > policy.max_cells:
                raise self._failure(
                    WorkbookScanFailureStatus.SCAN_LIMIT_EXCEEDED,
                    "WORKBOOK_CELL_LIMIT_EXCEEDED",
                    (
                        f"Workbook scan requires {total} cells, above the configured "
                        f"max_cells={policy.max_cells}."
                    ),
                    SourceLocation.range(sheet_name, sheet.calculate_dimension()),
                )
        return total

    def _scan_worksheet(
        self,
        formula_sheet: Worksheet,
        cached_sheet: Worksheet,
        *,
        position: int,
    ) -> SheetScan:
        sheet_name = str(formula_sheet.title)
        issues: list[ScanIssue] = []
        cells: list[CellEvidence] = []
        row_candidates: list[RowCandidate] = []
        prior_signatures: dict[tuple[str, ...], int] = {}
        merged_ranges = tuple(str(cell_range) for cell_range in formula_sheet.merged_cells.ranges)
        merged_row_intervals = tuple(
            (cell_range.min_row, cell_range.max_row)
            for cell_range in formula_sheet.merged_cells.ranges
        )

        for row in formula_sheet.iter_rows(
            min_row=1,
            max_row=formula_sheet.max_row,
            min_col=1,
            max_col=formula_sheet.max_column,
        ):
            row_values = [cell.value for cell in row]
            non_empty_values = [value for value in row_values if value is not None]
            row_index = int(row[0].row)
            signature = tuple(self._normalize_signature_value(value) for value in non_empty_values)

            if not non_empty_values:
                row_candidates.append(
                    RowCandidate(
                        row_index=row_index,
                        kind=RowCandidateKind.BLANK,
                        reason="No stored cell values in the used-column span.",
                    )
                )
            else:
                intersects_merge = any(
                    start <= row_index <= end for start, end in merged_row_intervals
                )
                if intersects_merge or (
                    len(non_empty_values) <= 2
                    and any(isinstance(value, str) for value in non_empty_values)
                ):
                    row_candidates.append(
                        RowCandidate(
                            row_index=row_index,
                            kind=RowCandidateKind.STRUCTURAL,
                            reason=(
                                "Row intersects a merged range."
                                if intersects_merge
                                else (
                                    "Sparse text row may be a section title, note, or "
                                    "signature row."
                                )
                            ),
                            signature=signature,
                        )
                    )

                if self._is_header_signature(signature, non_empty_values):
                    previous_row = prior_signatures.get(signature)
                    if previous_row is not None:
                        row_candidates.append(
                            RowCandidate(
                                row_index=row_index,
                                kind=RowCandidateKind.REPEATED_HEADER,
                                reason=(
                                    "Same header-like signature was first observed on row "
                                    f"{previous_row}."
                                ),
                                signature=signature,
                            )
                        )
                    else:
                        prior_signatures[signature] = row_index

            for cell in row:
                stored_value = cell.value
                if stored_value is None:
                    continue
                formula_text = (
                    str(stored_value)
                    if cell.data_type == "f"
                    or (isinstance(stored_value, str) and stored_value.startswith("="))
                    else None
                )
                cached_value = (
                    cached_sheet[cell.coordinate].value if formula_text is not None else None
                )
                evidence = CellEvidence(
                    coordinate=str(cell.coordinate),
                    stored_value=stored_value,
                    cached_value=cached_value,
                    formula_text=formula_text,
                    number_format=str(cell.number_format),
                    data_type=str(cell.data_type),
                )
                cells.append(evidence)
                issues.extend(self._cell_issues(sheet_name, evidence))

        protection = self._protection_metadata(formula_sheet)
        if protection.enabled:
            issues.append(
                ScanIssue(
                    code="SHEET_PROTECTION_OBSERVED",
                    severity=IssueSeverity.INFO,
                    message=(
                        "Protected sheet metadata was read without password, unlock, or "
                        "bypass attempts."
                    ),
                    location=SourceLocation.sheet(sheet_name),
                )
            )

        images = tuple(self._image_metadata(image) for image in formula_sheet._images)
        for image in images:
            location = (
                SourceLocation.cell(sheet_name, image.anchor_from)
                if image.anchor_from is not None
                else SourceLocation.sheet(sheet_name)
            )
            issues.append(
                ScanIssue(
                    code="IMAGE_METADATA_ONLY",
                    severity=IssueSeverity.INFO,
                    message=(
                        "Image presence and placement metadata were retained; content bytes "
                        "were not retained or analyzed."
                    ),
                    location=location,
                )
            )

        return SheetScan(
            name=sheet_name,
            kind=SheetKind.WORKSHEET,
            position=position,
            visibility=str(formula_sheet.sheet_state),
            used_range=formula_sheet.calculate_dimension(),
            estimated_cells=int(formula_sheet.max_row) * int(formula_sheet.max_column),
            merged_ranges=merged_ranges,
            hidden_row_ranges=self._hidden_row_ranges(formula_sheet),
            hidden_column_ranges=self._hidden_column_ranges(formula_sheet),
            cells=tuple(cells),
            row_candidates=tuple(row_candidates),
            protection=protection,
            images=images,
            issues=tuple(issues),
        )

    def _scan_chartsheet(self, sheet: Chartsheet, *, position: int) -> SheetScan:
        name = str(sheet.title)
        issue = ScanIssue(
            code="CHARTSHEET_STRUCTURE_ONLY",
            severity=IssueSeverity.INFO,
            message=(
                "Chart sheet existence, order, and visibility were retained; it has no "
                "worksheet cell grid."
            ),
            location=SourceLocation.sheet(name),
        )
        return SheetScan(
            name=name,
            kind=SheetKind.CHARTSHEET,
            position=position,
            visibility=str(sheet.sheet_state),
            used_range=None,
            estimated_cells=0,
            merged_ranges=(),
            hidden_row_ranges=(),
            hidden_column_ranges=(),
            cells=(),
            row_candidates=(),
            protection=SheetProtectionMetadata(enabled=False, protected_actions=()),
            images=(),
            issues=(issue,),
        )

    def _cell_issues(self, sheet_name: str, evidence: CellEvidence) -> list[ScanIssue]:
        issues: list[ScanIssue] = []
        refresh_reasons: list[str] = []
        location = SourceLocation.cell(sheet_name, evidence.coordinate)
        stored_text = str(evidence.stored_value)
        cached_text = "" if evidence.cached_value is None else str(evidence.cached_value)
        if evidence.formula_text is not None and _EXTERNAL_FORMULA_PATTERN.search(
            evidence.formula_text
        ):
            issues.append(
                ScanIssue(
                    code="EXTERNAL_REFERENCE_FORMULA",
                    severity=IssueSeverity.WARNING,
                    message=(
                        "Formula references an external workbook and must not be silently trusted."
                    ),
                    location=location,
                )
            )
            refresh_reasons.append("external workbook reference")
        formula_has_broken_reference = (
            evidence.formula_text is not None and "#REF!" in evidence.formula_text.upper()
        )
        stored_is_broken_reference_error = (
            evidence.data_type == "e" and stored_text.strip().upper() == "#REF!"
        )
        cached_formula_is_broken_reference_error = (
            evidence.formula_text is not None and cached_text.strip().upper() == "#REF!"
        )
        if (
            formula_has_broken_reference
            or stored_is_broken_reference_error
            or cached_formula_is_broken_reference_error
        ):
            issues.append(
                ScanIssue(
                    code="BROKEN_CELL_REFERENCE",
                    severity=IssueSeverity.WARNING,
                    message="Cell contains a broken #REF! reference.",
                    location=location,
                )
            )
            refresh_reasons.append("broken #REF! reference")
        if evidence.formula_text is not None and evidence.cached_value is None:
            issues.append(
                ScanIssue(
                    code="FORMULA_CACHE_MISSING",
                    severity=IssueSeverity.WARNING,
                    message=(
                        "Formula has no readable cached result; calculation refresh is "
                        "required before official use."
                    ),
                    location=location,
                )
            )
            refresh_reasons.append("missing formula cache")
        if refresh_reasons:
            issues.append(
                ScanIssue(
                    code="CALCULATION_REFRESH_REQUIRED",
                    severity=IssueSeverity.WARNING,
                    message=(
                        "Cell requires calculation refresh before official use: "
                        + ", ".join(refresh_reasons)
                        + "."
                    ),
                    location=location,
                )
            )
        return issues

    def _protection_metadata(self, sheet: Worksheet) -> SheetProtectionMetadata:
        protection = sheet.protection
        enabled = bool(protection.sheet)
        protected_actions = tuple(
            action for action in _PROTECTION_ACTIONS if bool(getattr(protection, action, False))
        )
        return SheetProtectionMetadata(
            enabled=enabled,
            protected_actions=protected_actions,
            password_material_collected=False,
            bypass_attempted=False,
        )

    def _hidden_row_ranges(self, sheet: Worksheet) -> tuple[IndexRange, ...]:
        indices = {
            int(index)
            for index, dimension in sheet.row_dimensions.items()
            if bool(dimension.hidden)
        }
        return self._compress_indices(indices)

    def _hidden_column_ranges(self, sheet: Worksheet) -> tuple[IndexRange, ...]:
        indices: set[int] = set()
        for key, dimension in sheet.column_dimensions.items():
            if not bool(dimension.hidden):
                continue
            start = int(dimension.min or column_index_from_string(str(key)))
            end = int(dimension.max or start)
            indices.update(range(start, end + 1))
        return self._compress_indices(indices)

    @staticmethod
    def _compress_indices(indices: set[int]) -> tuple[IndexRange, ...]:
        if not indices:
            return ()
        ordered = sorted(indices)
        ranges: list[IndexRange] = []
        start = previous = ordered[0]
        for current in ordered[1:]:
            if current == previous + 1:
                previous = current
                continue
            ranges.append(IndexRange(start=start, end=previous))
            start = previous = current
        ranges.append(IndexRange(start=start, end=previous))
        return tuple(ranges)

    @staticmethod
    def _normalize_signature_value(value: object) -> str:
        return " ".join(str(value).strip().casefold().split())

    @staticmethod
    def _is_header_signature(signature: tuple[str, ...], values: list[object]) -> bool:
        if len(signature) < 2:
            return False
        text_values = sum(isinstance(value, str) for value in values)
        return text_values >= max(2, len(values) // 2)

    def _image_metadata(self, image: Any) -> ImageMetadata:
        anchor = getattr(image, "anchor", None)
        anchor_from = self._anchor_coordinate(getattr(anchor, "_from", None))
        anchor_to = self._anchor_coordinate(getattr(anchor, "to", None))
        extent = getattr(anchor, "ext", None)
        return ImageMetadata(
            anchor_from=anchor_from,
            anchor_to=anchor_to,
            width_px=self._image_dimension(extent, "cx", getattr(image, "width", None)),
            height_px=self._image_dimension(extent, "cy", getattr(image, "height", None)),
            image_format=self._optional_string(getattr(image, "format", None)),
            content_collected=False,
            analysis_performed=False,
        )

    @staticmethod
    def _anchor_coordinate(marker: Any) -> str | None:
        if marker is None:
            return None
        column = int(marker.col) + 1
        row = int(marker.row) + 1
        return f"{get_column_letter(column)}{row}"

    @staticmethod
    def _safe_float(value: object) -> float | None:
        if value is None:
            return None
        try:
            return float(str(value))
        except (TypeError, ValueError):
            return None

    @classmethod
    def _image_dimension(
        cls,
        extent: object,
        extent_attribute: str,
        intrinsic_value: object,
    ) -> float | None:
        extent_value = getattr(extent, extent_attribute, None)
        if extent_value is not None:
            return float(EMU_to_pixels(int(extent_value)))
        return cls._safe_float(intrinsic_value)

    @staticmethod
    def _optional_string(value: object) -> str | None:
        return None if value is None else str(value)

    @staticmethod
    def _sha256(source: Path) -> str:
        digest = hashlib.sha256()
        try:
            with source.open("rb") as stream:
                for block in iter(lambda: stream.read(1024 * 1024), b""):
                    digest.update(block)
        except OSError as exc:
            raise OpenpyxlWorkbookScanner._failure(
                WorkbookScanFailureStatus.WORKBOOK_UNREADABLE,
                "WORKBOOK_HASH_READ_FAILED",
                "Workbook could not be read while computing its source hash.",
                SourceLocation.workbook(),
            ) from exc
        return digest.hexdigest()

    @classmethod
    def _sha256_stream(cls, source: BinaryIO) -> tuple[str, int]:
        original_position = cls._stream_position(source)
        digest = hashlib.sha256()
        size = 0
        try:
            source.seek(0)
            for block in iter(lambda: source.read(1024 * 1024), b""):
                digest.update(block)
                size += len(block)
        except (OSError, ValueError) as exc:
            raise cls._failure(
                WorkbookScanFailureStatus.WORKBOOK_UNREADABLE,
                "WORKBOOK_STREAM_HASH_READ_FAILED",
                "Seekable source stream could not be hashed.",
                SourceLocation.workbook(),
            ) from exc
        finally:
            try:
                source.seek(original_position)
            except (OSError, ValueError) as exc:
                raise cls._failure(
                    WorkbookScanFailureStatus.WORKBOOK_UNREADABLE,
                    "WORKBOOK_STREAM_POSITION_RESTORE_FAILED",
                    "Source stream position could not be restored after hashing.",
                    SourceLocation.workbook(),
                ) from exc
        return digest.hexdigest(), size

    @classmethod
    def _validated_stream_position(cls, source: BinaryIO) -> int:
        try:
            if not source.readable() or not source.seekable():
                raise cls._failure(
                    WorkbookScanFailureStatus.WORKBOOK_UNREADABLE,
                    "WORKBOOK_STREAM_NOT_SEEKABLE",
                    "File Store source must be a readable, seekable binary stream.",
                    SourceLocation.workbook(),
                )
            return int(source.tell())
        except WorkbookScanFailure:
            raise
        except (AttributeError, OSError, ValueError) as exc:
            raise cls._failure(
                WorkbookScanFailureStatus.WORKBOOK_UNREADABLE,
                "WORKBOOK_STREAM_NOT_SEEKABLE",
                "File Store source must be a readable, seekable binary stream.",
                SourceLocation.workbook(),
            ) from exc

    @staticmethod
    def _stream_position(source: BinaryIO) -> int:
        try:
            return int(source.tell())
        except (OSError, ValueError) as exc:
            raise OpenpyxlWorkbookScanner._failure(
                WorkbookScanFailureStatus.WORKBOOK_UNREADABLE,
                "WORKBOOK_STREAM_POSITION_UNAVAILABLE",
                "Source stream position is unavailable.",
                SourceLocation.workbook(),
            ) from exc

    @classmethod
    def _sha256_or_none(cls, source: Path) -> str | None:
        try:
            return cls._sha256(source)
        except WorkbookScanFailure:
            return None

    def _raise_with_integrity_evidence(
        self,
        source: Path,
        before_hash: str,
        failure: WorkbookScanFailure,
    ) -> Never:
        after_hash = self._sha256_or_none(source)
        if after_hash != before_hash:
            raise self._source_mutated_failure(before_hash, after_hash) from failure
        raise WorkbookScanFailure(
            failure.status,
            failure.issue,
            source_sha256_before=before_hash,
            source_sha256_after=after_hash,
        ) from failure.__cause__

    def _raise_with_stream_integrity_evidence(
        self,
        source: BinaryIO,
        before_hash: str,
        failure: WorkbookScanFailure,
    ) -> Never:
        try:
            after_hash, _ = self._sha256_stream(source)
        except WorkbookScanFailure:
            after_hash = None
        if after_hash != before_hash:
            raise self._source_mutated_failure(before_hash, after_hash) from failure
        raise WorkbookScanFailure(
            failure.status,
            failure.issue,
            source_sha256_before=before_hash,
            source_sha256_after=after_hash,
        ) from failure.__cause__

    @staticmethod
    def _source_mutated_failure(
        before_hash: str,
        after_hash: str | None,
    ) -> WorkbookScanFailure:
        return WorkbookScanFailure(
            WorkbookScanFailureStatus.SOURCE_MUTATED_DURING_SCAN,
            ScanIssue(
                code="SOURCE_MUTATED_DURING_SCAN",
                severity=IssueSeverity.ERROR,
                message="Source workbook changed or became unreadable while it was being scanned.",
                location=SourceLocation.workbook(),
            ),
            source_sha256_before=before_hash,
            source_sha256_after=after_hash,
        )

    @staticmethod
    def _failure(
        status: WorkbookScanFailureStatus,
        code: str,
        message: str,
        location: SourceLocation,
    ) -> WorkbookScanFailure:
        return WorkbookScanFailure(
            status,
            ScanIssue(
                code=code,
                severity=IssueSeverity.ERROR,
                message=message,
                location=location,
            ),
        )
