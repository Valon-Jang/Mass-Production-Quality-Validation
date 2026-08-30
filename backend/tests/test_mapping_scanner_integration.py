from __future__ import annotations

import hashlib
from datetime import UTC, date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook  # type: ignore[import-untyped]

from app.application.mapping_preview import (
    InMemoryMappingTemplateRegistry,
    build_mapping_preview,
)
from app.domain.mapping import (
    CellAddress,
    HeaderTokenAssertion,
    IdentifierKind,
    IdentifierMapping,
    InspectionRowMapping,
    MappingPreviewRequest,
    MappingPreviewState,
    MappingTemplate,
    MappingTemplateStatus,
    MergeSignatureAssertion,
    RowStructureAssertion,
    SheetStructureAssertion,
    SystemJudgmentStatus,
    WorkbookFingerprint,
)
from app.domain.workbook_scan import SheetKind
from app.infrastructure.excel import OpenpyxlWorkbookScanner

_REPORT = "Generic Report"
_RAW = "Raw Data"
_REFERENCE = "Reference"


def _address(sheet_name: str, coordinate: str) -> CellAddress:
    return CellAddress(sheet_name=sheet_name, coordinate=coordinate)


def _save_synthetic_workbook(path: Path) -> None:
    workbook = Workbook()
    report = workbook.active
    report.title = _REPORT
    report.merge_cells("A1:L1")
    report["A1"] = "Generic Quality Report"
    report["A2"] = "Supplier"
    report["B2"] = "SUPPLIER-GENERIC"
    report["C2"] = "Model"
    report["D2"] = "MODEL-GENERIC"
    report["E2"] = "Lot"
    report["F2"] = "LOT-GENERIC"
    report["G2"] = "Inspection Date"
    report["H2"] = "2026-06-01"
    headers = (
        "Item",
        "Method",
        "Instrument",
        "Specification",
        "Tolerance",
        "Minimum",
        "Maximum",
        "Sample 1",
        "Sample 2",
        "Sample 3",
        "Sample 4",
        "Supplier Result",
    )
    for column, header_value in enumerate(headers, start=1):
        report.cell(row=3, column=column, value=header_value)

    numeric_row = (
        "Length characteristic",
        "Comparator method",
        "Instrument class",
        "Nominal requirement",
        "Symmetric tolerance",
        9.5,
        "=F4+1",
        10.0,
        10.1,
        9.9,
        10.0,
        "OK",
    )
    for column, numeric_value in enumerate(numeric_row, start=1):
        report.cell(row=4, column=column, value=numeric_value)
    for coordinate in ("F4", "G4", "H4", "I4", "J4", "K4"):
        report[coordinate].number_format = "0.00"

    qualitative_row = (
        "Surface condition",
        "Visual method",
        "Inspection station",
        "Qualitative requirement",
        "Not numeric",
        "N/A",
        "N/A",
        "Clear",
        "Clear",
        "OK",
    )
    for column, qualitative_value in enumerate(qualitative_row, start=1):
        report.cell(row=5, column=column, value=qualitative_value)

    raw = workbook.create_sheet(_RAW)
    raw.sheet_state = "hidden"
    raw["A1"] = "Raw section"
    raw.append(("Source", "Sequence", "Status"))

    reference = workbook.create_sheet(_REFERENCE)
    reference.sheet_state = "veryHidden"
    reference.merge_cells("A1:B1")
    reference["A1"] = "Reference section"
    reference.append(("Generic code", "Generic description"))
    workbook.save(path)


def _row_mapping(
    row_key: str,
    row_number: int,
    sample_columns: tuple[str, ...],
    result_column: str,
) -> InspectionRowMapping:
    return InspectionRowMapping(
        row_key=row_key,
        item=_address(_REPORT, f"A{row_number}"),
        method=_address(_REPORT, f"B{row_number}"),
        instrument=_address(_REPORT, f"C{row_number}"),
        specification=_address(_REPORT, f"D{row_number}"),
        tolerance=_address(_REPORT, f"E{row_number}"),
        minimum=_address(_REPORT, f"F{row_number}"),
        maximum=_address(_REPORT, f"G{row_number}"),
        sample_cells=tuple(_address(_REPORT, f"{column}{row_number}") for column in sample_columns),
        supplier_result=_address(_REPORT, f"{result_column}{row_number}"),
    )


def _template() -> MappingTemplate:
    numeric_row = _row_mapping("numeric-row", 4, ("H", "I", "J", "K"), "L")
    qualitative_row = _row_mapping("qualitative-row", 5, ("H", "I"), "J")
    return MappingTemplate(
        template_id="synthetic-three-sheet-template",
        schema_version="1",
        revision=1,
        status=MappingTemplateStatus.APPROVED,
        project_key="integration-project",
        supplier_scope="integration-supplier-scope",
        supplier_source_aliases=("SUPPLIER-GENERIC",),
        approved_by="integration-reviewer",
        approved_at=datetime(2026, 1, 1, tzinfo=UTC),
        effective_from=date(2026, 1, 1),
        effective_to=date(2026, 12, 31),
        fingerprint=WorkbookFingerprint(
            header_tokens=(
                HeaderTokenAssertion(
                    _address(_REPORT, "A1"),
                    "Generic Quality Report",
                ),
                HeaderTokenAssertion(_address(_REPORT, "A2"), "Supplier"),
                HeaderTokenAssertion(_address(_REPORT, "C2"), "Model"),
                HeaderTokenAssertion(_address(_REPORT, "E2"), "Lot"),
                HeaderTokenAssertion(_address(_REPORT, "G2"), "Inspection Date"),
                HeaderTokenAssertion(_address(_REPORT, "A3"), "Item"),
                HeaderTokenAssertion(_address(_REPORT, "L3"), "Supplier Result"),
                HeaderTokenAssertion(_address(_RAW, "A1"), "Raw section"),
                HeaderTokenAssertion(
                    _address(_REFERENCE, "A1"),
                    "Reference section",
                ),
            ),
            sheet_structures=(
                SheetStructureAssertion(
                    sheet_name=_REPORT,
                    expected_position=0,
                    expected_kind=SheetKind.WORKSHEET,
                    expected_visibility="visible",
                    expected_used_range="A1:L5",
                ),
                SheetStructureAssertion(
                    sheet_name=_RAW,
                    expected_position=1,
                    expected_kind=SheetKind.WORKSHEET,
                    expected_visibility="hidden",
                    expected_used_range="A1:C2",
                ),
                SheetStructureAssertion(
                    sheet_name=_REFERENCE,
                    expected_position=2,
                    expected_kind=SheetKind.WORKSHEET,
                    expected_visibility="veryHidden",
                    expected_used_range="A1:B2",
                ),
            ),
            merge_signatures=(
                MergeSignatureAssertion(_REPORT, ("A1:L1",)),
                MergeSignatureAssertion(_RAW, ()),
                MergeSignatureAssertion(_REFERENCE, ("A1:B1",)),
            ),
            row_structures=(
                RowStructureAssertion(
                    row_key=numeric_row.row_key,
                    sheet_name=_REPORT,
                    row_index=4,
                    expected_non_empty_cells=numeric_row.all_addresses,
                ),
                RowStructureAssertion(
                    row_key=qualitative_row.row_key,
                    sheet_name=_REPORT,
                    row_index=5,
                    expected_non_empty_cells=qualitative_row.all_addresses,
                ),
            ),
        ),
        identifiers=(
            IdentifierMapping(IdentifierKind.SUPPLIER, _address(_REPORT, "B2")),
            IdentifierMapping(IdentifierKind.MODEL, _address(_REPORT, "D2")),
            IdentifierMapping(IdentifierKind.LOT_NUMBER, _address(_REPORT, "F2")),
            IdentifierMapping(IdentifierKind.INSPECTION_DATE, _address(_REPORT, "H2")),
        ),
        inspection_rows=(numeric_row, qualitative_row),
    )


@pytest.mark.required_test_id("DQ-P1-MAP-008")
def test_scanner_output_flows_into_mapping_preview_without_official_judgment(
    tmp_path: Path,
) -> None:
    source = tmp_path / "synthetic-integration.xlsx"
    _save_synthetic_workbook(source)
    expected_hash = hashlib.sha256(source.read_bytes()).hexdigest()

    scan = OpenpyxlWorkbookScanner().scan(source)
    source_issues = scan.issues
    registry = InMemoryMappingTemplateRegistry()
    registry.register(_template())
    result = build_mapping_preview(
        scan,
        MappingPreviewRequest(
            project_key="integration-project",
            supplier_scope="integration-supplier-scope",
        ),
        registry,
    )

    assert scan.source_sha256_before == expected_hash
    assert scan.source_sha256_after == expected_hash
    assert scan.source_sha256_before
    assert hashlib.sha256(source.read_bytes()).hexdigest() == expected_hash
    assert scan.issues == source_issues
    assert {issue.code for issue in source_issues}.issuperset(
        {"FORMULA_CACHE_MISSING", "CALCULATION_REFRESH_REQUIRED"}
    )

    assert result.state == MappingPreviewState.PREVIEW_READY
    assert result.preview is not None
    assert result.issues == ()
    assert result.preview.source_sha256_before == expected_hash
    assert result.preview.source_sha256_after == expected_hash
    assert result.preview.source_sha256_before
    assert result.preview.is_golden_workbook_evidence is False
    assert result.preview.source_issues == source_issues
    identifiers = {item.kind: item.evidence for item in result.preview.identifiers}
    assert identifiers[IdentifierKind.SUPPLIER].source == _address(_REPORT, "B2")
    assert identifiers[IdentifierKind.MODEL].raw_value == "MODEL-GENERIC"
    assert identifiers[IdentifierKind.LOT_NUMBER].source == _address(_REPORT, "F2")
    assert identifiers[IdentifierKind.INSPECTION_DATE].raw_value == "2026-06-01"

    numeric_row, qualitative_row = result.preview.inspection_rows
    assert numeric_row.item.source == _address(_REPORT, "A4")
    assert numeric_row.item.raw_value == "Length characteristic"
    assert [sample.raw_value for sample in numeric_row.samples] == [10.0, 10.1, 9.9, 10.0]
    assert numeric_row.maximum is not None
    assert numeric_row.maximum.source == _address(_REPORT, "G4")
    assert numeric_row.maximum.formula_text == "=F4+1"
    assert numeric_row.maximum.cached_value is None
    assert numeric_row.specification is not None
    assert numeric_row.supplier_result is not None
    assert numeric_row.specification.source == _address(_REPORT, "D4")
    assert numeric_row.supplier_result.source == _address(_REPORT, "L4")
    assert numeric_row.supplier_result.raw_value == "OK"
    assert qualitative_row.samples[0].source == _address(_REPORT, "H5")
    assert qualitative_row.samples[0].raw_value == "Clear"
    assert qualitative_row.supplier_result is not None
    assert qualitative_row.supplier_result.source == _address(_REPORT, "J5")
    assert all(
        row.system_judgment_status == SystemJudgmentStatus.NOT_EVALUATED
        and row.system_judgment is None
        for row in result.preview.inspection_rows
    )
    assert result.official_values_created is False
    assert result.calculations_performed is False
