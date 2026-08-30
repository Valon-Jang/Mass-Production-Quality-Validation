from __future__ import annotations

import hashlib
from dataclasses import replace
from datetime import UTC, date, datetime
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook  # type: ignore[import-untyped]

from app.application.long_candidate import LongCandidateInputError, build_long_candidate
from app.application.manual_ingestion import ManualWorkbookIngestionService
from app.application.mapping_preview import (
    InMemoryMappingTemplateRegistry,
    MappingTemplateCatalog,
)
from app.application.mapping_template_commands import (
    ApproveMappingTemplateRevisionCommand,
    CreateMappingTemplateRevisionCommand,
    MappingTemplateCommandService,
    ReviewMappingTemplateRevisionCommand,
)
from app.application.store_scan_mapping import (
    ResolvedMappingScope,
    StoreScanMappingOutcome,
    StoreScanMappingRequest,
    StoreScanMappingService,
    StoreScanMappingStatus,
)
from app.domain.identity import Actor, ActorKind, Role
from app.domain.long_format import (
    CanonicalRowBinding,
    CanonicalRowBindingCatalog,
    CanonicalRowBindingKey,
    CanonicalRowBindingStatus,
    LongCandidateState,
    LongDataStatus,
    LongIssueCode,
    LongRowState,
    MaterializedCanonicalRowBindingCatalog,
    MeasurementMode,
    SamplePolicy,
    SpecEvaluationStatus,
    UnitConversionStatus,
)
from app.domain.mapping import (
    CellAddress,
    HeaderTokenAssertion,
    IdentifierKind,
    IdentifierMapping,
    InspectionRowMapping,
    MappingIssue,
    MappingIssueCode,
    MappingPreview,
    MappingPreviewResult,
    MappingPreviewState,
    MappingTemplate,
    MappingTemplateStatus,
    MergeSignatureAssertion,
    RowStructureAssertion,
    SheetStructureAssertion,
    SystemJudgmentStatus,
    WorkbookFingerprint,
)
from app.domain.workbook_scan import (
    DisplayValueStatus,
    ScanPolicy,
    SheetKind,
    SourceLocationKind,
)
from app.infrastructure.database import Base, Database
from app.infrastructure.excel import OpenpyxlWorkbookScanner
from app.infrastructure.file_store import XLSX_MIME, OriginalFileStore
from app.infrastructure.mapping_templates import MappingTemplateRepository

_SHEET = "Synthetic Report"
_PROJECT = "project-synthetic"
_SUPPLIER_SCOPE = "supplier-scope-synthetic"
_SOURCE_SUPPLIER = "SUPPLIER-SYNTHETIC"
_SOURCE_MODEL = "MODEL-SYNTHETIC"
_SOURCE_LOT = "LOT-SYNTHETIC"
_INSPECTION_DATE = date(2026, 6, 15)
_APPROVAL_TIME = datetime(2026, 1, 2, 3, 4, tzinfo=UTC)
_WORKFLOW_TIME = datetime(2026, 1, 3, 4, 5, tzinfo=UTC)
_SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

_REVIEWER = Actor(
    actor_id="synthetic-reviewer",
    kind=ActorKind.LOCAL_OWNER,
    roles=frozenset({Role.REVIEWER}),
)
_ADMIN = Actor(
    actor_id="synthetic-admin",
    kind=ActorKind.LOCAL_OWNER,
    roles=frozenset({Role.ADMIN}),
)


def _address(coordinate: str) -> CellAddress:
    return CellAddress(sheet_name=_SHEET, coordinate=coordinate)


def _row_mapping(
    row_key: str,
    row_number: int,
    sample_columns: tuple[str, ...],
) -> InspectionRowMapping:
    return InspectionRowMapping(
        row_key=row_key,
        item=_address(f"A{row_number}"),
        method=_address(f"B{row_number}"),
        instrument=_address(f"C{row_number}"),
        specification=_address(f"D{row_number}"),
        tolerance=_address(f"E{row_number}"),
        minimum=_address(f"F{row_number}"),
        maximum=_address(f"G{row_number}"),
        sample_cells=tuple(_address(f"{column}{row_number}") for column in sample_columns),
        supplier_result=_address(f"J{row_number}"),
    )


def _template(
    *,
    include_zero_row: bool = False,
    status: MappingTemplateStatus = MappingTemplateStatus.APPROVED,
) -> MappingTemplate:
    rows = [
        _row_mapping("numeric-row", 4, ("H", "I")),
        _row_mapping("qualitative-row", 5, ("H",)),
    ]
    if include_zero_row:
        rows.append(_row_mapping("zero-sample-row", 6, ()))
    expected_used_range = "A1:N6" if include_zero_row else "A1:N5"
    approved = status == MappingTemplateStatus.APPROVED
    return MappingTemplate(
        template_id="synthetic-long-template",
        schema_version="1",
        revision=1,
        status=status,
        project_key=_PROJECT,
        supplier_scope=_SUPPLIER_SCOPE,
        supplier_source_aliases=(_SOURCE_SUPPLIER,),
        approved_by="synthetic-approver" if approved else None,
        approved_at=_APPROVAL_TIME if approved else None,
        effective_from=date(2026, 1, 1),
        effective_to=date(2026, 12, 31),
        fingerprint=WorkbookFingerprint(
            header_tokens=(
                HeaderTokenAssertion(_address("A1"), "Synthetic Long Candidate"),
                HeaderTokenAssertion(_address("A2"), "Supplier"),
                HeaderTokenAssertion(_address("C2"), "Model"),
                HeaderTokenAssertion(_address("E2"), "Lot"),
                HeaderTokenAssertion(_address("G2"), "Inspection Date"),
                HeaderTokenAssertion(_address("I2"), "Part"),
                HeaderTokenAssertion(_address("K2"), "Report"),
                HeaderTokenAssertion(_address("M2"), "Revision"),
                HeaderTokenAssertion(_address("A3"), "Item"),
                HeaderTokenAssertion(_address("H3"), "Sample 1"),
                HeaderTokenAssertion(_address("J3"), "Supplier Result"),
            ),
            sheet_structures=(
                SheetStructureAssertion(
                    sheet_name=_SHEET,
                    expected_position=0,
                    expected_kind=SheetKind.WORKSHEET,
                    expected_visibility="visible",
                    expected_used_range=expected_used_range,
                ),
            ),
            merge_signatures=(MergeSignatureAssertion(_SHEET, ()),),
            row_structures=tuple(
                RowStructureAssertion(
                    row_key=row.row_key,
                    sheet_name=_SHEET,
                    row_index=row.item.row_index,
                    expected_non_empty_cells=row.all_addresses,
                )
                for row in rows
            ),
        ),
        identifiers=(
            IdentifierMapping(IdentifierKind.SUPPLIER, _address("B2")),
            IdentifierMapping(IdentifierKind.MODEL, _address("D2")),
            IdentifierMapping(IdentifierKind.LOT_NUMBER, _address("F2")),
            IdentifierMapping(IdentifierKind.INSPECTION_DATE, _address("H2")),
            IdentifierMapping(IdentifierKind.PART_NUMBER, _address("J2")),
            IdentifierMapping(IdentifierKind.REPORT_NUMBER, _address("L2")),
            IdentifierMapping(IdentifierKind.REVISION, _address("N2")),
        ),
        inspection_rows=tuple(rows),
    )


def _save_workbook(
    path: Path,
    *,
    include_zero_row: bool = False,
    formula_samples: bool = False,
    qualitative_sample: str = "CLEAR",
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _SHEET
    sheet["A1"] = "Synthetic Long Candidate"
    metadata = {
        "A2": "Supplier",
        "B2": _SOURCE_SUPPLIER,
        "C2": "Model",
        "D2": _SOURCE_MODEL,
        "E2": "Lot",
        "F2": _SOURCE_LOT,
        "G2": "Inspection Date",
        "H2": _INSPECTION_DATE,
        "I2": "Part",
        "J2": "PART-SYNTHETIC",
        "K2": "Report",
        "L2": "REPORT-SYNTHETIC",
        "M2": "Revision",
        "N2": "REV-SYNTHETIC",
    }
    for coordinate, value in metadata.items():
        sheet[coordinate] = value
    headers = (
        "Item",
        "Method",
        "Instrument",
        "Specification",
        "Tolerance",
        "Minimum",
        "Maximum",
        "Sample 1",
        "Sample 2",
        "Supplier Result",
    )
    for column, value in enumerate(headers, start=1):
        sheet.cell(row=3, column=column, value=value)
    numeric_values: tuple[object, ...] = (
        "Synthetic dimension",
        "Synthetic comparator",
        "Synthetic gauge",
        "Synthetic nominal requirement",
        "Synthetic tolerance",
        9.0,
        11.0,
        "=F4+1.25" if formula_samples else 10.25,
        "=F4+0.75" if formula_samples else 9.75,
        "SUPPLIER-OK",
    )
    qualitative_values: tuple[object, ...] = (
        "Synthetic appearance",
        "Synthetic visual method",
        "Synthetic station",
        "Synthetic qualitative requirement",
        "Synthetic qualitative tolerance",
        "N/A",
        "N/A",
        qualitative_sample,
        None,
        "SUPPLIER-OK",
    )
    for row_number, values in ((4, numeric_values), (5, qualitative_values)):
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row_number, column=column, value=value)
    if include_zero_row:
        zero_values: tuple[object, ...] = (
            "Synthetic document check",
            "Synthetic review method",
            "Synthetic record",
            "Synthetic record requirement",
            "N/A",
            "N/A",
            "N/A",
            None,
            None,
            "SUPPLIER-OK",
        )
        for column, value in enumerate(zero_values, start=1):
            sheet.cell(row=6, column=column, value=value)
    workbook.save(path)
    if formula_samples:
        _inject_formula_cache(path, coordinate="H4", value="10.25")
    return path.read_bytes()


def _inject_formula_cache(path: Path, *, coordinate: str, value: str) -> None:
    temporary = path.with_suffix(".cached.tmp")
    with (
        ZipFile(path, "r") as source_archive,
        ZipFile(temporary, "w", compression=ZIP_DEFLATED) as target_archive,
    ):
        for info in source_archive.infolist():
            content = source_archive.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                root = ElementTree.fromstring(content)
                cell = next(
                    candidate
                    for candidate in root.findall(f".//{{{_SHEET_NS}}}c")
                    if candidate.attrib.get("r") == coordinate
                )
                cached = cell.find(f"{{{_SHEET_NS}}}v")
                assert cached is not None
                cached.text = value
                content = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            target_archive.writestr(info, content)
    temporary.replace(path)


def _execute_route(
    tmp_path: Path,
    *,
    name: str,
    mapping_catalog: MappingTemplateCatalog,
    include_zero_row: bool = False,
    formula_samples: bool = False,
    qualitative_sample: str = "CLEAR",
    model_candidates: tuple[str, ...] = (_SOURCE_MODEL,),
    lot_candidates: tuple[str, ...] = (_SOURCE_LOT,),
) -> tuple[StoreScanMappingOutcome, Path, bytes]:
    tmp_path.mkdir(parents=True, exist_ok=True)
    source = tmp_path / f"{name}.xlsx"
    original = _save_workbook(
        source,
        include_zero_row=include_zero_row,
        formula_samples=formula_samples,
        qualitative_sample=qualitative_sample,
    )
    store = OriginalFileStore(tmp_path / "s", max_bytes=1024 * 1024)
    route = StoreScanMappingService(
        ingestion_service=ManualWorkbookIngestionService(
            file_store=store,
            scanner=OpenpyxlWorkbookScanner(),
        ),
        registry=mapping_catalog,
    )
    outcome = route.execute(
        StoreScanMappingRequest(
            scope=ResolvedMappingScope(
                project_key=_PROJECT,
                supplier_scope=_SUPPLIER_SCOPE,
            ),
            source=source,
            declared_mime_type=XLSX_MIME,
            scan_policy=ScanPolicy(max_cells=10_000),
            model_candidates=model_candidates,
            lot_candidates=lot_candidates,
        )
    )
    assert outcome.status == StoreScanMappingStatus.PREVIEW_READY
    return outcome, source, original


def _ready_outcome(
    tmp_path: Path,
    *,
    name: str,
    template: MappingTemplate,
    include_zero_row: bool = False,
    formula_samples: bool = False,
    qualitative_sample: str = "CLEAR",
    model_candidates: tuple[str, ...] = (_SOURCE_MODEL,),
    lot_candidates: tuple[str, ...] = (_SOURCE_LOT,),
) -> tuple[StoreScanMappingOutcome, Path, bytes]:
    registry = InMemoryMappingTemplateRegistry()
    registry.register(template)
    return _execute_route(
        tmp_path,
        name=name,
        mapping_catalog=registry,
        include_zero_row=include_zero_row,
        formula_samples=formula_samples,
        qualitative_sample=qualitative_sample,
        model_candidates=model_candidates,
        lot_candidates=lot_candidates,
    )


def _binding(
    template: MappingTemplate,
    row_key: str,
    *,
    binding_revision: int = 1,
    source_model_values: tuple[str, ...] = (_SOURCE_MODEL,),
    canonical_model_key: str = "canonical-model-synthetic",
    canonical_supplier_key: str = "canonical-supplier-synthetic",
    sample_policy: SamplePolicy = SamplePolicy.AT_LEAST_ONE,
    measurement_mode: MeasurementMode | None = None,
    status: CanonicalRowBindingStatus = CanonicalRowBindingStatus.APPROVED,
    effective_from: date = date(2026, 1, 1),
    effective_to: date | None = date(2026, 12, 31),
) -> CanonicalRowBinding:
    if measurement_mode is None:
        if row_key == "numeric-row":
            measurement_mode = MeasurementMode.NUMERIC
        elif row_key == "qualitative-row":
            measurement_mode = MeasurementMode.QUALITATIVE
        else:
            measurement_mode = MeasurementMode.JUDGMENT_ONLY
    approved = status == CanonicalRowBindingStatus.APPROVED
    return CanonicalRowBinding(
        key=CanonicalRowBindingKey(
            project_key=template.project_key,
            supplier_scope=template.supplier_scope,
            template_id=template.template_id,
            template_revision=template.revision,
            row_key=row_key,
        ),
        binding_revision=binding_revision,
        status=status,
        approved_by="synthetic-binding-approver" if approved else None,
        approved_at=_APPROVAL_TIME if approved else None,
        effective_from=effective_from,
        effective_to=effective_to,
        source_model_values=source_model_values,
        canonical_model_key=canonical_model_key,
        canonical_supplier_key=canonical_supplier_key,
        canonical_model_part_key="canonical-part-synthetic",
        canonical_item_key=f"canonical-item-{row_key}",
        sample_policy=sample_policy,
        measurement_mode=measurement_mode,
    )


def _catalog(
    template: MappingTemplate,
    *,
    zero_policy: SamplePolicy = SamplePolicy.ZERO_ALLOWED,
    zero_mode: MeasurementMode = MeasurementMode.JUDGMENT_ONLY,
) -> MaterializedCanonicalRowBindingCatalog:
    return MaterializedCanonicalRowBindingCatalog(
        bindings=tuple(
            _binding(
                template,
                row.row_key,
                sample_policy=(
                    zero_policy if row.row_key == "zero-sample-row" else SamplePolicy.AT_LEAST_ONE
                ),
                measurement_mode=zero_mode if row.row_key == "zero-sample-row" else None,
            )
            for row in template.inspection_rows
        )
    )


def _preview(outcome: StoreScanMappingOutcome) -> MappingPreview:
    assert outcome.mapping_result is not None
    assert outcome.mapping_result.preview is not None
    return outcome.mapping_result.preview


@pytest.mark.required_test_id("DQ-P1-LONG-001")
def test_ready_candidate_preserves_exact_provenance_and_mixed_variable_samples(
    tmp_path: Path,
) -> None:
    template = _template()
    outcome, source, original = _ready_outcome(
        tmp_path,
        name="ready",
        template=template,
    )
    preview = _preview(outcome)
    catalog = _catalog(template)

    result = build_long_candidate(outcome, catalog)

    assert result == build_long_candidate(outcome, catalog)
    assert result.state == LongCandidateState.LOAD_CANDIDATE_READY
    assert len(result.rows) == 2
    assert result.loadable_rows == result.rows
    assert result.held_rows == ()
    assert result.official_values_created is False
    assert result.calculations_performed is False
    provenance = result.provenance
    expected_hash = hashlib.sha256(original).hexdigest()
    assert provenance.receipt is outcome.receipt
    assert provenance.receipt.received_at == outcome.receipt.received_at
    assert provenance.receipt.model_candidates == (_SOURCE_MODEL,)
    assert provenance.receipt.lot_candidates == (_SOURCE_LOT,)
    assert (
        provenance.receipt.content_sha256
        == provenance.preview_sha256_before
        == provenance.preview_sha256_after
        == expected_hash
    )
    assert provenance.preview_source_name == source.name
    assert provenance.preview_source_size_bytes == len(original)
    assert provenance.source_issues == preview.source_issues
    assert provenance.is_golden_workbook_evidence is False
    assert provenance.template_id == template.template_id
    assert provenance.template_schema_version == template.schema_version
    assert provenance.template_revision == template.revision
    assert provenance.template_approved_by == template.approved_by
    assert provenance.template_approved_at == template.approved_at
    assert provenance.template_effective_from == template.effective_from
    assert provenance.template_effective_to == template.effective_to
    assert provenance.source_inspection_date == _INSPECTION_DATE
    assert provenance.binding_catalog_revision == catalog.catalog_revision
    assert provenance.binding_catalog_revision.startswith("sha256:")
    assert tuple(
        selection.matches[0].binding_revision for selection in provenance.binding_selections
    ) == (1, 1)
    binding_signatures = tuple(selection.matches[0] for selection in provenance.binding_selections)
    assert tuple(signature.status for signature in binding_signatures) == (
        CanonicalRowBindingStatus.APPROVED,
        CanonicalRowBindingStatus.APPROVED,
    )
    assert tuple(signature.measurement_mode for signature in binding_signatures) == (
        MeasurementMode.NUMERIC,
        MeasurementMode.QUALITATIVE,
    )
    assert all(
        signature.approved_by == "synthetic-binding-approver" for signature in binding_signatures
    )
    assert all(signature.approved_at == _APPROVAL_TIME for signature in binding_signatures)
    assert all(signature.effective_from == date(2026, 1, 1) for signature in binding_signatures)
    assert all(signature.effective_to == date(2026, 12, 31) for signature in binding_signatures)
    assert result.source_identifiers == preview.identifiers
    assert {identifier.kind for identifier in result.source_identifiers} == {
        IdentifierKind.MODEL,
        IdentifierKind.PART_NUMBER,
        IdentifierKind.LOT_NUMBER,
        IdentifierKind.SUPPLIER,
        IdentifierKind.INSPECTION_DATE,
        IdentifierKind.REPORT_NUMBER,
        IdentifierKind.REVISION,
    }
    assert result.source_model is not None
    assert result.source_model.raw_value == _SOURCE_MODEL
    assert result.source_lot is not None
    assert result.source_lot.raw_value == _SOURCE_LOT

    numeric_row, qualitative_row = result.rows
    assert numeric_row.data_status == LongDataStatus.PENDING
    assert numeric_row.spec_evaluation_status == SpecEvaluationStatus.NOT_EVALUATED
    assert numeric_row.system_judgment_status == SystemJudgmentStatus.NOT_EVALUATED
    assert numeric_row.system_judgment is None
    assert numeric_row.specification is not None
    assert numeric_row.specification.raw_value == "Synthetic nominal requirement"
    assert numeric_row.supplier_judgment is not None
    assert numeric_row.supplier_judgment.raw_value == "SUPPLIER-OK"
    assert tuple(sample.raw_numeric_value for sample in numeric_row.measurements) == (10.25, 9.75)
    assert all(sample.raw_qualitative_value is None for sample in numeric_row.measurements)
    assert tuple(sample.evidence.source.coordinate for sample in numeric_row.measurements) == (
        "H4",
        "I4",
    )
    assert qualitative_row.measurements[0].raw_numeric_value is None
    assert qualitative_row.measurements[0].raw_qualitative_value == "CLEAR"
    for row in result.rows:
        assert row.system_judgment_status == SystemJudgmentStatus.NOT_EVALUATED
        assert row.system_judgment is None
        for sample in row.measurements:
            assert sample.standardized_value is None
            assert sample.unit_conversion_status == UnitConversionStatus.NOT_CONFIGURED
            assert sample.evidence.cached_value is None
            assert sample.evidence.formula_text is None
            assert sample.evidence.display_value is None
            assert sample.evidence.display_value_status == DisplayValueStatus.NOT_RENDERED

    changed_bindings = tuple(
        replace(binding, binding_revision=2) if binding.key.row_key == "numeric-row" else binding
        for binding in catalog.bindings
    )
    changed_catalog = MaterializedCanonicalRowBindingCatalog(bindings=changed_bindings)
    changed_result = build_long_candidate(outcome, changed_catalog)
    assert changed_catalog.catalog_revision != catalog.catalog_revision
    assert changed_result.provenance.deterministic_key != provenance.deterministic_key
    assert source.read_bytes() == original


@pytest.mark.required_test_id("DQ-P1-LONG-002")
def test_zero_samples_require_explicit_policy_without_dropping_row_evidence(
    tmp_path: Path,
) -> None:
    template = _template(include_zero_row=True)
    outcome, source, original = _ready_outcome(
        tmp_path,
        name="zero-sample",
        template=template,
        include_zero_row=True,
    )
    preview = _preview(outcome)
    allowed = build_long_candidate(
        outcome,
        _catalog(template),
    )

    assert allowed.state == LongCandidateState.LOAD_CANDIDATE_READY
    zero_row = next(row for row in allowed.rows if row.row_key == "zero-sample-row")
    assert zero_row.measurements == ()
    assert zero_row.binding is not None
    assert zero_row.binding.sample_policy == SamplePolicy.ZERO_ALLOWED
    assert zero_row.binding.measurement_mode == MeasurementMode.JUDGMENT_ONLY
    assert zero_row.item == preview.inspection_rows[2].item
    assert zero_row.supplier_judgment == preview.inspection_rows[2].supplier_result

    restricted = build_long_candidate(
        outcome,
        _catalog(
            template,
            zero_policy=SamplePolicy.AT_LEAST_ONE,
            zero_mode=MeasurementMode.NUMERIC,
        ),
    )
    assert restricted.state == LongCandidateState.PARTIAL_HOLD
    restricted_zero = next(row for row in restricted.rows if row.row_key == "zero-sample-row")
    assert restricted_zero.state == LongRowState.ROW_HELD
    assert restricted_zero.measurements == ()
    assert restricted_zero.item == zero_row.item
    assert restricted_zero.method == zero_row.method
    assert restricted_zero.instrument == zero_row.instrument
    assert restricted_zero.specification == zero_row.specification
    assert restricted_zero.supplier_judgment == zero_row.supplier_judgment
    assert {issue.code for issue in restricted_zero.issues} == {
        LongIssueCode.ZERO_SAMPLE_POLICY_REQUIRED
    }
    assert all(row.data_status == LongDataStatus.PENDING for row in restricted.rows)
    with pytest.raises(ValueError, match="JUDGMENT_ONLY"):
        _binding(
            template,
            "zero-sample-row",
            sample_policy=SamplePolicy.AT_LEAST_ONE,
            measurement_mode=MeasurementMode.JUDGMENT_ONLY,
        )

    blank_outcome, _, _ = _ready_outcome(
        tmp_path / "q",
        name="q",
        template=_template(),
        qualitative_sample="   ",
    )
    blank_result = build_long_candidate(blank_outcome, _catalog(_template()))
    blank_row = next(row for row in blank_result.rows if row.row_key == "qualitative-row")
    assert blank_result.state == LongCandidateState.PARTIAL_HOLD
    assert blank_row.measurements[0].raw_qualitative_value is None
    assert {issue.code for issue in blank_row.issues} == {LongIssueCode.INVALID_QUALITATIVE_SAMPLE}
    assert source.read_bytes() == original


@pytest.mark.required_test_id("DQ-P1-LONG-003")
def test_missing_ambiguous_and_non_ready_inputs_fail_or_hold_explicitly(
    tmp_path: Path,
) -> None:
    template = _template()
    outcome, _, _ = _ready_outcome(tmp_path, name="binding-holds", template=template)
    all_bindings = _catalog(template).bindings

    partial = build_long_candidate(
        outcome,
        MaterializedCanonicalRowBindingCatalog(bindings=(all_bindings[0],)),
    )
    assert partial.state == LongCandidateState.PARTIAL_HOLD
    assert partial.rows[0].state == LongRowState.LOADABLE_PENDING
    assert partial.rows[1].state == LongRowState.ROW_HELD
    assert partial.rows[1].issues[0].code == LongIssueCode.CANONICAL_ROW_BINDING_MISSING

    held = build_long_candidate(
        outcome,
        MaterializedCanonicalRowBindingCatalog(bindings=()),
    )
    assert held.state == LongCandidateState.LOAD_HELD
    assert len(held.rows) == len(_preview(outcome).inspection_rows)
    assert all(row.state == LongRowState.ROW_HELD for row in held.rows)

    ambiguous_binding = replace(all_bindings[0], binding_revision=2)
    ambiguous_catalog = MaterializedCanonicalRowBindingCatalog(
        bindings=(all_bindings[0], ambiguous_binding, all_bindings[1])
    )
    ambiguous = build_long_candidate(outcome, ambiguous_catalog)
    assert ambiguous.state == LongCandidateState.PARTIAL_HOLD
    assert ambiguous.rows[0].binding is None
    assert ambiguous.rows[0].issues[0].code == LongIssueCode.CANONICAL_ROW_BINDING_AMBIGUOUS
    assert len(ambiguous.provenance.binding_selections[0].matches) == 2
    with pytest.raises(ValueError, match="revision identity"):
        MaterializedCanonicalRowBindingCatalog(bindings=(all_bindings[0], all_bindings[0]))
    with pytest.raises(ValueError, match="held row"):
        replace(partial.rows[0], state=LongRowState.ROW_HELD)

    ready = build_long_candidate(outcome, _catalog(template))
    with pytest.raises(ValueError, match="loadable row"):
        replace(
            partial.rows[1],
            state=LongRowState.LOADABLE_PENDING,
            binding=all_bindings[1],
        )
    with pytest.raises(ValueError, match="ready candidate"):
        replace(ready, issues=(partial.issues[0],))
    with pytest.raises(ValueError, match="hold evidence"):
        replace(
            ready,
            state=LongCandidateState.LOAD_HELD,
            rows=(),
            issues=(),
        )

    draft_catalog = MaterializedCanonicalRowBindingCatalog(
        bindings=tuple(
            replace(
                binding,
                status=CanonicalRowBindingStatus.DRAFT,
                approved_by=None,
                approved_at=None,
            )
            for binding in all_bindings
        )
    )
    draft_result = build_long_candidate(outcome, draft_catalog)
    assert draft_result.state == LongCandidateState.LOAD_HELD
    assert {issue.code for issue in draft_result.issues} == {
        LongIssueCode.CANONICAL_ROW_BINDING_NOT_APPROVED
    }

    future_catalog = MaterializedCanonicalRowBindingCatalog(
        bindings=tuple(
            replace(
                binding,
                effective_from=date(2027, 1, 1),
                effective_to=date(2027, 12, 31),
            )
            for binding in all_bindings
        )
    )
    future_result = build_long_candidate(outcome, future_catalog)
    assert future_result.state == LongCandidateState.LOAD_HELD
    assert {issue.code for issue in future_result.issues} == {
        LongIssueCode.CANONICAL_ROW_BINDING_NOT_EFFECTIVE
    }

    expired_catalog = MaterializedCanonicalRowBindingCatalog(
        bindings=tuple(
            replace(
                binding,
                effective_from=date(2025, 1, 1),
                effective_to=date(2025, 12, 31),
            )
            for binding in all_bindings
        )
    )
    expired_result = build_long_candidate(outcome, expired_catalog)
    assert expired_result.state == LongCandidateState.LOAD_HELD
    assert {issue.code for issue in expired_result.issues} == {
        LongIssueCode.CANONICAL_ROW_BINDING_NOT_EFFECTIVE
    }
    assert (
        len(
            {
                _catalog(template).catalog_revision,
                draft_catalog.catalog_revision,
                future_catalog.catalog_revision,
                expired_catalog.catalog_revision,
            }
        )
        == 4
    )

    mapping_required = StoreScanMappingOutcome(
        status=StoreScanMappingStatus.RAW_PRESERVED_MAPPING_REQUIRED,
        scope=outcome.scope,
        receipt=outcome.receipt,
        scan=outcome.scan,
        mapping_result=MappingPreviewResult(
            state=MappingPreviewState.MAPPING_REQUIRED,
            preview=None,
            issues=(
                MappingIssue(
                    code=MappingIssueCode.TEMPLATE_MISSING,
                    message="Synthetic Mapping Template is intentionally unavailable.",
                ),
            ),
        ),
    )
    with pytest.raises(LongCandidateInputError, match="PREVIEW_READY"):
        build_long_candidate(mapping_required, _catalog(template))


class _WrongScopeCatalog:
    catalog_revision = "synthetic-wrong-scope-v1"

    def find(self, key: CanonicalRowBindingKey) -> tuple[CanonicalRowBinding, ...]:
        template = _template()
        binding = _binding(template, key.row_key)
        return (
            replace(
                binding,
                key=replace(
                    key,
                    project_key="different-project",
                    supplier_scope="different-supplier",
                ),
            ),
        )


@pytest.mark.required_test_id("DQ-P1-LONG-004")
def test_scope_model_and_lot_conflicts_globally_hold_without_losing_evidence(
    tmp_path: Path,
) -> None:
    template = _template()
    outcome, _, _ = _ready_outcome(
        tmp_path,
        name="candidate-conflict",
        template=template,
        model_candidates=("DIFFERENT-MODEL",),
        lot_candidates=("DIFFERENT-LOT",),
    )
    preview = _preview(outcome)

    globally_held = build_long_candidate(outcome, _catalog(template))

    assert globally_held.state == LongCandidateState.LOAD_HELD
    assert {issue.code for issue in globally_held.issues}.issuperset(
        {LongIssueCode.MODEL_CANDIDATE_CONFLICT, LongIssueCode.LOT_CANDIDATE_CONFLICT}
    )
    assert globally_held.source_identifiers == preview.identifiers
    for candidate_row, preview_row in zip(
        globally_held.rows,
        preview.inspection_rows,
        strict=True,
    ):
        assert candidate_row.state == LongRowState.ROW_HELD
        assert candidate_row.binding is not None
        assert candidate_row.item == preview_row.item
        assert candidate_row.method == preview_row.method
        assert candidate_row.instrument == preview_row.instrument
        assert candidate_row.specification == preview_row.specification
        assert candidate_row.tolerance == preview_row.tolerance
        assert candidate_row.minimum == preview_row.minimum
        assert candidate_row.maximum == preview_row.maximum
        assert (
            tuple(sample.evidence for sample in candidate_row.measurements) == preview_row.samples
        )
        assert candidate_row.supplier_judgment == preview_row.supplier_result

    model_mismatch_catalog = MaterializedCanonicalRowBindingCatalog(
        bindings=tuple(
            replace(binding, source_model_values=("DIFFERENT-MODEL",))
            for binding in _catalog(template).bindings
        )
    )
    model_mismatch = build_long_candidate(outcome, model_mismatch_catalog)
    assert all(
        LongIssueCode.SOURCE_MODEL_BINDING_CONFLICT in {issue.code for issue in row.issues}
        for row in model_mismatch.rows
    )

    wrong_scope_catalog: CanonicalRowBindingCatalog = _WrongScopeCatalog()
    wrong_scope = build_long_candidate(outcome, wrong_scope_catalog)
    wrong_scope_codes = {issue.code for issue in wrong_scope.issues}
    assert LongIssueCode.PROJECT_SCOPE_CONFLICT in wrong_scope_codes
    assert LongIssueCode.SUPPLIER_SCOPE_CONFLICT in wrong_scope_codes
    assert wrong_scope.state == LongCandidateState.LOAD_HELD


@pytest.mark.required_test_id("DQ-P1-LONG-005")
def test_formula_cache_and_refresh_are_preserved_but_never_projected_as_raw(
    tmp_path: Path,
) -> None:
    template = _template()
    outcome, source, original = _ready_outcome(
        tmp_path,
        name="formula-evidence",
        template=template,
        formula_samples=True,
    )
    preview = _preview(outcome)

    result = build_long_candidate(outcome, _catalog(template))

    assert result.state == LongCandidateState.PARTIAL_HOLD
    formula_row = result.rows[0]
    assert formula_row.state == LongRowState.ROW_HELD
    cached_formula, refresh_formula = formula_row.measurements
    assert cached_formula.evidence.source.coordinate == "H4"
    assert cached_formula.evidence.raw_value == "=F4+1.25"
    assert cached_formula.evidence.formula_text == "=F4+1.25"
    assert cached_formula.evidence.cached_value == 10.25
    assert cached_formula.raw_numeric_value is None
    assert cached_formula.raw_qualitative_value is None
    assert refresh_formula.evidence.source.coordinate == "I4"
    assert refresh_formula.evidence.raw_value == "=F4+0.75"
    assert refresh_formula.evidence.formula_text == "=F4+0.75"
    assert refresh_formula.evidence.cached_value is None
    assert refresh_formula.raw_numeric_value is None
    assert refresh_formula.raw_qualitative_value is None
    formula_issue_locations = {
        (issue.row_key, issue.sheet_name, issue.coordinate)
        for issue in formula_row.issues
        if issue.code == LongIssueCode.FORMULA_SAMPLE_NOT_RAW
    }
    assert formula_issue_locations == {
        ("numeric-row", _SHEET, "H4"),
        ("numeric-row", _SHEET, "I4"),
    }
    mode_issue_locations = {
        (issue.row_key, issue.sheet_name, issue.coordinate)
        for issue in formula_row.issues
        if issue.code == LongIssueCode.MEASUREMENT_MODE_MISMATCH
    }
    assert mode_issue_locations == {
        ("numeric-row", _SHEET, "H4"),
        ("numeric-row", _SHEET, "I4"),
    }
    refresh_issue_locations = {
        (issue.row_key, issue.sheet_name, issue.coordinate)
        for issue in formula_row.issues
        if issue.code == LongIssueCode.CALCULATION_REFRESH_REQUIRED
    }
    assert refresh_issue_locations == {("numeric-row", _SHEET, "I4")}
    scanner_refresh_locations = {
        (issue.location.kind, issue.location.sheet_name, issue.location.coordinate)
        for issue in preview.source_issues
        if issue.code == "CALCULATION_REFRESH_REQUIRED"
    }
    assert scanner_refresh_locations == {(SourceLocationKind.CELL, _SHEET, "I4")}
    scan = outcome.scan
    assert scan is not None
    assert result.provenance.source_issues == preview.source_issues == scan.issues
    assert formula_row.supplier_judgment is not None
    assert formula_row.supplier_judgment.raw_value == "SUPPLIER-OK"
    assert formula_row.system_judgment_status == SystemJudgmentStatus.NOT_EVALUATED
    assert formula_row.system_judgment is None
    assert result.official_values_created is False
    assert result.calculations_performed is False
    assert source.read_bytes() == original

    normal_outcome, _, _ = _ready_outcome(
        tmp_path / "m",
        name="m",
        template=template,
    )
    normal_bindings = _catalog(template).bindings
    wrong_mode_catalog = MaterializedCanonicalRowBindingCatalog(
        bindings=(
            replace(normal_bindings[0], measurement_mode=MeasurementMode.QUALITATIVE),
            normal_bindings[1],
        )
    )
    wrong_mode = build_long_candidate(normal_outcome, wrong_mode_catalog)
    assert wrong_mode.state == LongCandidateState.PARTIAL_HOLD
    wrong_mode_row = wrong_mode.rows[0]
    assert all(sample.raw_numeric_value is None for sample in wrong_mode_row.measurements)
    assert {(issue.code, issue.coordinate) for issue in wrong_mode_row.issues} == {
        (LongIssueCode.MEASUREMENT_MODE_MISMATCH, "H4"),
        (LongIssueCode.MEASUREMENT_MODE_MISMATCH, "I4"),
    }


@pytest.mark.required_test_id("DQ-P1-LONG-006")
def test_persisted_approved_store_scan_mapping_to_candidate_is_pure_and_pending(
    tmp_path: Path,
) -> None:
    database_path = tmp_path / "mapping.sqlite3"
    database = Database(f"sqlite+pysqlite:///{database_path.as_posix()}")
    Base.metadata.create_all(database.engine)
    repository = MappingTemplateRepository()
    commands = MappingTemplateCommandService(
        database,
        repository=repository,
        clock=lambda: _WORKFLOW_TIME,
    )
    draft = _template(status=MappingTemplateStatus.DRAFT)
    try:
        created = commands.create_revision(
            CreateMappingTemplateRevisionCommand(
                template=draft,
                expected_history_row_version=0,
                actor=_REVIEWER,
                reason="Create a synthetic Long candidate Mapping Template.",
                source_reference="synthetic-long-fixture",
            )
        )
        reviewed = commands.review(
            ReviewMappingTemplateRevisionCommand(
                project_key=_PROJECT,
                supplier_scope=_SUPPLIER_SCOPE,
                template_id=draft.template_id,
                revision=draft.revision,
                expected_history_row_version=created.history_row_version,
                expected_revision_row_version=created.revision_row_version,
                actor=_REVIEWER,
                reason="Review every synthetic source-cell assertion.",
            )
        )
        approved = commands.approve(
            ApproveMappingTemplateRevisionCommand(
                project_key=_PROJECT,
                supplier_scope=_SUPPLIER_SCOPE,
                template_id=draft.template_id,
                revision=draft.revision,
                expected_history_row_version=reviewed.history_row_version,
                expected_revision_row_version=reviewed.revision_row_version,
                actor=_ADMIN,
                reason="Approve the reviewed synthetic mapping revision.",
            )
        )
        assert approved.template.status == MappingTemplateStatus.APPROVED
        with database.session() as session:
            persistent_catalog: MappingTemplateCatalog = repository.load_catalog(
                session,
                project_key=_PROJECT,
            )
    finally:
        database.dispose()

    database_before = database_path.read_bytes()
    outcome, source, original = _execute_route(
        tmp_path,
        name="persistent-approved",
        mapping_catalog=persistent_catalog,
    )
    candidate = build_long_candidate(outcome, _catalog(approved.template))

    assert outcome.status == StoreScanMappingStatus.PREVIEW_READY
    assert candidate.state == LongCandidateState.LOAD_CANDIDATE_READY
    assert all(row.data_status == LongDataStatus.PENDING for row in candidate.rows)
    assert all(row.system_judgment is None for row in candidate.rows)
    assert candidate.provenance.template_approved_by == _ADMIN.actor_id
    assert candidate.provenance.template_approved_at == _WORKFLOW_TIME
    assert candidate.official_values_created is False
    assert candidate.calculations_performed is False
    assert database_path.read_bytes() == database_before
    assert source.read_bytes() == original
